"""
Top brands report (private — not published to the site).

Reads the per-month partner summary that build_enriched_monthly.py already
produces and writes a formatted Excel report. The site only ever shows the top
10 brands; this exposes the full ranking.

Usage:
    python scripts/report_top_brands.py                  # latest month, top 200
    python scripts/report_top_brands.py --month 2026-07
    python scripts/report_top_brands.py --top 50
    python scripts/report_top_brands.py --top 0          # every brand

Output:
    reports/top-<N>-brands-<month>.xlsx
"""

import argparse
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parent

DATA_DIR = PROJECT_ROOT / "data"
MONTHLY_INDEX_FILE = DATA_DIR / "monthly" / "index.json"
DERIVED_ROOT = DATA_DIR / "derived" / "monthly"
REPORTS_DIR = PROJECT_ROOT / "reports"

NOTHS_PARTNER_URL = "https://www.notonthehighstreet.com/partners/{slug}"

# NOTHS purple, matching the site.
HEADER_FILL = PatternFill("solid", fgColor="7066E0")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
LINK_FONT = Font(color="0563C1", underline="single")


def load_json(path: Path):
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def latest_month() -> str:
    """Most recent month present in the monthly index."""
    if not MONTHLY_INDEX_FILE.exists():
        raise FileNotFoundError(f"Monthly index not found: {MONTHLY_INDEX_FILE}")

    months = [
        m["month"]
        for m in load_json(MONTHLY_INDEX_FILE).get("months", [])
        if isinstance(m, dict) and m.get("month")
    ]
    if not months:
        raise ValueError("No months found in the monthly index.")

    return max(months)


def competition_ranks(rows: list[dict]) -> None:
    """
    Assign 1, 2, 2, 4 style ranks in place — brands on equal reviews share a
    rank, matching how the site presents ties.
    """
    previous_reviews = None
    previous_rank = 0

    for i, row in enumerate(rows, start=1):
        reviews = row["total_reviews_month"]
        if reviews == previous_reviews:
            row["rank"] = previous_rank
        else:
            row["rank"] = i
            previous_rank = i
            previous_reviews = reviews


def build_report(month: str, top_n: int) -> Path:
    partners_file = DERIVED_ROOT / month / "partners_summary.json"
    summary_file = DERIVED_ROOT / month / "summary.json"

    if not partners_file.exists():
        raise FileNotFoundError(
            f"No partner summary for {month}: {partners_file}\n"
            f"Run scripts/build_enriched_monthly.py first."
        )

    partners = load_json(partners_file)
    summary = load_json(summary_file) if summary_file.exists() else {}

    # partners_summary.json is already sorted, but don't rely on it.
    partners.sort(
        key=lambda p: (
            -(p.get("total_reviews_month") or 0),
            (p.get("seller_name") or "").lower(),
        )
    )

    total_brands = len(partners)
    rows = partners if top_n <= 0 else partners[:top_n]
    competition_ranks(rows)

    month_reviews = summary.get("total_reviews_month") or sum(
        (p.get("total_reviews_month") or 0) for p in partners
    )

    wb = Workbook()
    ws = wb.active
    ws.title = f"Top brands {month}"

    headers = [
        "Rank",
        "Brand",
        "Slug",
        "Reviews",
        "Products",
        "Avg reviews per product",
        "Share of month",
        "NOTHS page",
    ]
    ws.append(headers)

    for row in rows:
        reviews = row.get("total_reviews_month") or 0
        products = row.get("product_count_month") or 0
        slug = row.get("seller_slug") or ""

        ws.append([
            row["rank"],
            row.get("seller_name") or slug,
            slug,
            reviews,
            products,
            round(reviews / products, 2) if products else 0,
            (reviews / month_reviews) if month_reviews else 0,
            NOTHS_PARTNER_URL.format(slug=slug) if slug else "",
        ])

    # --- formatting ---------------------------------------------------------
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"

    for r in range(2, ws.max_row + 1):
        ws.cell(row=r, column=6).number_format = "0.00"
        ws.cell(row=r, column=7).number_format = "0.0%"

        link_cell = ws.cell(row=r, column=8)
        if link_cell.value:
            link_cell.hyperlink = link_cell.value
            link_cell.font = LINK_FONT

    widths = [7, 38, 26, 10, 11, 13, 13, 52]
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.row_dimensions[1].height = 30

    # --- context sheet ------------------------------------------------------
    meta = wb.create_sheet("About")
    for line in [
        ("Report", f"Top brands by reviews received during {month}"),
        ("Brands listed", len(rows)),
        ("Brands with reviews", total_brands),
        ("Total reviews in month", month_reviews),
        ("Products without a resolved brand", summary.get("products_without_seller", "n/a")),
        ("Reviews without a resolved brand", summary.get("reviews_without_seller", "n/a")),
        ("Source", f"data/derived/monthly/{month}/partners_summary.json"),
        ("Generated by", "scripts/report_top_brands.py"),
    ]:
        meta.append(list(line))

    for cell in meta["A"]:
        cell.font = Font(bold=True)
    meta.column_dimensions["A"].width = 36
    meta.column_dimensions["B"].width = 56

    REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    label = "all" if top_n <= 0 else str(top_n)
    out_path = REPORTS_DIR / f"top-{label}-brands-{month}.xlsx"
    wb.save(out_path)

    return out_path


def main() -> None:
    parser = argparse.ArgumentParser(description="Build a top brands Excel report.")
    parser.add_argument("--month", help="Month to report on, e.g. 2026-07. Defaults to the latest.")
    parser.add_argument("--top", type=int, default=200, help="How many brands to list. 0 for all.")
    args = parser.parse_args()

    month = args.month or latest_month()
    out_path = build_report(month, args.top)

    print(f"✅ {month}: report written → {out_path}")


if __name__ == "__main__":
    main()
